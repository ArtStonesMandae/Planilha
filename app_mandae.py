
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io
import re

st.set_page_config(page_title="Planilha Mandae", layout="centered")

# Fundo personalizado na área central do app
st.markdown("""
    <style>
    .stApp {
        background-color: #FBE9E7;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("<h2 style='text-align:center; color:#333;'>Gerador de Planilhas Mandae</h2>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center; color:#555;'>Suba o arquivo CSV dos pedidos e gere sua planilha formatada com 1 clique!</p>", unsafe_allow_html=True)

uploaded_file = st.file_uploader("📎 Selecione o arquivo CSV", type=["csv"])

if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file, encoding='latin1', sep=';', dtype=str)
    except Exception as ex:
        st.error(f"Erro ao ler o arquivo: {ex}")
        st.stop()

    if df['Destinatário'].isna().any():
        st.error("⚠️ Existem linhas com DESTINATÁRIO vazio. Corrija antes de continuar.")
        st.stop()

    def format_document(cpf, cnpj):
        if pd.notna(cnpj):
            return re.sub(r'\D', '', cnpj).zfill(14)
        elif pd.notna(cpf):
            return re.sub(r'\D', '', cpf).zfill(11)
        return ''

    def get_phone(row):
        return row['Telefone'] if pd.notna(row['Telefone']) else row['Celular']

    def clean_cep(cep):
        return re.sub(r'\D', '', cep)

    saida_df = pd.DataFrame({
        'NOME DO DESTINATÁRIO*': df['Destinatário'],
        'NOME DA EMPRESA (EM CASO DE ENDEREÇO COMERCIAL)': df['Razão Social'],
        'E-MAIL': df['Email'],
        'TELEFONE': df.apply(get_phone, axis=1),
        'CPF / CNPJ CLIENTE*': df.apply(lambda row: format_document(row['CPF'], row['CNPJ']), axis=1),
        'INSCR. ESTADUAL': df['Inscrição Estadual'],
        'CEP*': df['Cep'].apply(clean_cep),
        'LOGRADOURO*': df['Endereço'],
        'NÚMERO*': df['Número'],
        'COMPLEMENTO': df['Complemento'],
        'BAIRRO*': df['Bairro'],
        'CIDADE*': df['Cidade'],
        'ESTADO*': df['Estado'],
        'PONTO DE REFERÊNCIA': '',
        'VOLUMES*': '1',
        'A ENCOMENDA POSSUI NF?*': 'Sim',
        'CHAVE NF': df['Nome do Cliente'],
        'CÓDIGO INTERNO DA SUA EMPRESA (OPCIONAL)': df['Pedido'],
        'SERVIÇO DE ENVIO*': df['Frete tipo'],
        'QR CODE (Não utilizar)': '',
        'VALOR DECLARADO (OPCIONAL)': df['Subtotal produtos'],
        'OBSERVAÇÃO': df['Obs. cliente']
    })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha Mandae"

    roxo_escuro = "FF5F497A"
    rosa_claro = "FFFBE9E7"
    texto_roxo_escuro = "FF5F497A"
    branco = "FFFFFFFF"
    cinza_sem_borda = ['N', 'O', 'P', 'T', 'V']
    bordas = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fonte_titulo_branco = Font(color=branco, bold=True)
    fonte_roxo = Font(color=texto_roxo_escuro)
    alinhado_centro = Alignment(horizontal="center", vertical="center")
    alinhado_esquerda = Alignment(horizontal="left", vertical="center")
    alinhado_direita = Alignment(horizontal="right", vertical="center")

    ws.merge_cells('A1:F1')
    ws.merge_cells('G1:N1')
    ws.merge_cells('O1:R1')
    ws.merge_cells('S1:V1')
    for col in range(1, 23):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color=roxo_escuro, end_color=roxo_escuro, fill_type='solid')
        cell.font = fonte_titulo_branco
        cell.alignment = alinhado_centro
        cell.border = bordas
    ws['A1'] = "DESTINATÁRIO"
    ws['G1'] = "ENDEREÇO"
    ws['O1'] = "ENCOMENDA"
    ws['S1'] = "PEDIDO MANDAE"

    ws.append(list(saida_df.columns))
    for cell in ws[2]:
        cell.fill = PatternFill(start_color=rosa_claro, end_color=rosa_claro, fill_type='solid')
        cell.font = fonte_roxo
        cell.alignment = alinhado_centro
    ws.row_dimensions[2].height = 30

    larguras_personalizadas = {
        'A': 37.14, 'B': 30.71, 'C': 13.00, 'D': 15.71,
        'G': 12.85, 'Q': 51.42, 'V': 15.71
    }
    for col_idx in range(1, 23):
        col_letter = get_column_letter(col_idx)
        largura = larguras_personalizadas.get(col_letter, 20)
        ws.column_dimensions[col_letter].width = largura

    for i, row in saida_df.iterrows():
        ws.append(row.tolist())

    for row in ws.iter_rows(min_row=3, max_row=2+len(saida_df), min_col=1, max_col=22):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in cinza_sem_borda:
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
            else:
                cell.fill = PatternFill(start_color=rosa_claro, end_color=rosa_claro, fill_type='solid')
                cell.border = bordas
            cell.alignment = alinhado_direita if col_letter == 'O' else alinhado_esquerda

    wb.save(output)
    output.seek(0)

    hoje = datetime.today()
    dia_util = hoje + timedelta(days=1)
    if hoje.weekday() == 4:
        dia_util += timedelta(days=2)
    nome_arquivo = f"{len(saida_df)}Pedidos - {dia_util.strftime('%d.%m')} - L2.xlsx"

    st.markdown("<p style='color:#333333; font-weight:500;'>✅ Sua planilha tá prontinha! Baixe no botão abaixo:</p>", unsafe_allow_html=True)
    st.download_button(label="📥 Baixar Planilha", data=output, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
